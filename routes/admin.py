from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from functools import wraps
from models import db, User, Document
from utils import ROLE_CHOICES, log_activity, generate_doc_id
import secrets
import string
from datetime import datetime, date
import os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'admin':
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@admin_bp.route('/users')
@login_required
@admin_required
def users_list():
    users = User.query.all()
    return render_template('admin/users.html', users=users)

@admin_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        role = request.form.get('role', 'consultant_limited')

        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('admin.add_user'))

        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('admin.add_user'))

        password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))

        user = User(username=username, email=email, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        log_activity(current_user.id, 'CREATE_USER', 'user', user.id, f'Role: {role}')

        flash(f'User {username} created successfully. Password: {password}', 'success')
        return redirect(url_for('admin.users_list'))

    return render_template('admin/add_user.html', roles=ROLE_CHOICES)

@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        user.email = request.form.get('email')
        user.role = request.form.get('role')

        if user.role == 'consultant_limited':
            selected_doc_ids = request.form.getlist('assigned_documents')
            user.assigned_documents = Document.query.filter(Document.doc_id.in_(selected_doc_ids)).all()

        db.session.commit()

        log_activity(current_user.id, 'UPDATE_USER', 'user', user_id)

        flash(f'User {user.username} updated successfully.', 'success')
        return redirect(url_for('admin.users_list'))

    documents = Document.query.all()
    return render_template('admin/edit_user.html', user=user, documents=documents, roles=ROLE_CHOICES)

@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    if user_id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.users_list'))

    user = User.query.get_or_404(user_id)
    username = user.username
    db.session.delete(user)
    db.session.commit()

    log_activity(current_user.id, 'DELETE_USER', 'user', user_id, f'Username: {username}')

    flash(f'User {username} deleted successfully.', 'success')
    return redirect(url_for('admin.users_list'))

@admin_bp.route('/import-excel', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        if not openpyxl:
            flash('Excel import requires openpyxl library. Please install it.', 'danger')
            return redirect(url_for('admin.import_excel'))

        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('admin.import_excel'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('admin.import_excel'))

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    header_key = cell.value.lower().strip()
                    headers[header_key] = col_idx

            def find_header(headers, *keywords):
                for keyword in keywords:
                    keyword_lower = keyword.lower()
                    for header_key, col_idx in headers.items():
                        if keyword_lower in header_key or header_key in keyword_lower:
                            return col_idx
                return None

            doc_id_col = find_header(headers, 'doc id', 'document id', 'doc_id')
            doc_name_col = find_header(headers, 'document name', 'name', 'doc name')

            if not doc_id_col or not doc_name_col:
                flash('Excel file must contain "Document ID" and "Document Name" columns.', 'danger')
                return redirect(url_for('admin.import_excel'))

            imported = 0
            skipped = 0
            errors = []

            format_col = find_header(headers, 'format')
            description_col = find_header(headers, 'description', 'desc')
            status_col = find_header(headers, 'status')
            distribution_col = find_header(headers, 'distribution', 'dist')
            owner_col = find_header(headers, 'owner', 'owner party')

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                try:
                    doc_id = row[doc_id_col - 1].value
                    doc_name = row[doc_name_col - 1].value

                    if not doc_id or not doc_name:
                        skipped += 1
                        continue

                    if Document.query.filter_by(doc_id=str(doc_id)).first():
                        skipped += 1
                        continue

                    doc_format = row[format_col - 1].value if format_col else '.docx'
                    description = row[description_col - 1].value if description_col else ''
                    status = row[status_col - 1].value if status_col else 'Draft'
                    distribution = row[distribution_col - 1].value if distribution_col else 'Internal'
                    owner_party = row[owner_col - 1].value if owner_col else 'Internal'

                    doc = Document(
                        doc_id=str(doc_id),
                        name=str(doc_name),
                        format=doc_format or '.docx',
                        description=description or '',
                        status=status or 'Draft',
                        distribution=distribution or 'Internal',
                        owner_party=owner_party or 'Internal',
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    db.session.add(doc)
                    imported += 1
                except Exception as e:
                    errors.append(f'Row {row_idx}: {str(e)}')
                    continue

            db.session.commit()

            log_activity(current_user.id, 'IMPORT_EXCEL', 'document', 0, f'Imported: {imported}, Skipped: {skipped}')

            flash(f'Import completed! {imported} documents imported, {skipped} skipped.', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:5])}', 'warning')

            return redirect(url_for('documents.documents_list'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'danger')
            return redirect(url_for('admin.import_excel'))

    return render_template('admin/import_excel.html')

@admin_bp.route('/download-template')
def download_template():
    if not openpyxl:
        flash('Template generation requires openpyxl library.', 'danger')
        return redirect(url_for('admin.import_excel'))

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Documents'

        headers = ['Document ID', 'Document Name', 'Format', 'Status', 'Distribution', 'Owner', 'Description']
        ws.append(headers)

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sample_data = [
            ['DOC-01', 'SPV Structuring Proposal', '.docx', 'Tracked — file not held', 'Internal', 'Internal', 'Generic five-party SPV model'],
            ['DOC-02', 'Salt Plant SPV & Sukuk Proposal', '.docx', 'Tracked — file not held', 'Internal', 'Internal', 'SPV model applied to project'],
            ['DOC-06', 'Mutual NDA — AASB', '.docx', 'Active — Pending Signature', 'External (AASB)', 'Calmic', 'Mutual NDA; AASB as Project Consultant'],
        ]

        for row_data in sample_data:
            ws.append(row_data)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40

        template_path = '/tmp/MDR_Template.xlsx'
        wb.save(template_path)

        log_activity(current_user.id, 'DOWNLOAD_TEMPLATE', 'template', 0)

        return send_file(
            template_path,
            as_attachment=True,
            download_name='MDR_Template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'danger')
        return redirect(url_for('admin.import_excel'))
